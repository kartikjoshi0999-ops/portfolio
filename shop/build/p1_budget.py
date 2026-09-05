import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from kjstyle import *

OUT = os.path.join(os.path.dirname(__file__), "..", "dist")
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
CATS = [
    ("Income","Employment income"),("Income","Self-employment / side income"),
    ("Income","Government benefits (CCB, GST/HST credit)"),("Income","Investment / other income"),
    ("Housing","Rent or mortgage payment"),("Housing","Property tax"),
    ("Housing","Home / tenant insurance"),("Housing","Utilities (hydro, gas, water)"),
    ("Housing","Internet / phone"),("Housing","Maintenance & repairs"),
    ("Living","Groceries"),("Living","Dining out & takeout"),("Living","Household supplies"),
    ("Transport","Car payment"),("Transport","Fuel"),("Transport","Auto insurance"),
    ("Transport","Repairs, plates & licence"),("Transport","Transit / rideshare"),
    ("Health","Health & dental premiums"),("Health","Prescriptions & medical"),
    ("Health","Fitness"),
    ("Family","Childcare"),("Family","School & activities"),("Family","Pets"),
    ("Debt","Credit card payments"),("Debt","Student loan payments"),("Debt","Line of credit / loan"),
    ("Savings","TFSA contribution"),("Savings","RRSP contribution"),
    ("Savings","RESP contribution"),("Savings","Emergency fund"),
    ("Lifestyle","Subscriptions & streaming"),("Lifestyle","Clothing & personal care"),
    ("Lifestyle","Travel & holidays"),("Lifestyle","Gifts & donations"),("Lifestyle","Entertainment"),
    ("Other","Bank fees"),("Other","Miscellaneous"),
]
SPARES = 12                      # blank slots already inside every formula range
SLOTS = len(CATS) + SPARES
FIRST = 5
LAST = FIRST + SLOTS - 1

wb = Workbook(); wb.remove(wb.active)

# ---------- Categories ----------
cat = wb.create_sheet("Categories")
title_block(cat, "Categories", "Rename any row to match your life, or fill in the spare rows at the bottom.", 4)
header_row(cat, 4, ["Group", "Category", "Type"])
for k in range(SLOTS):
    r = FIRST + k
    if k < len(CATS):
        g, c = CATS[k]
        cat.cell(row=r, column=1, value=g).border = BOX
        input_cell(cat, f"B{r}", c)
        t = "Income" if g == "Income" else ("Savings" if g == "Savings" else "Expense")
        cat.cell(row=r, column=3, value=t).border = BOX
    else:
        input_cell(cat, f"A{r}")
        input_cell(cat, f"B{r}")
        input_cell(cat, f"C{r}")
widths(cat, {"A": 16, "B": 40, "C": 14})
note(cat, LAST + 2, f"The last {SPARES} rows are deliberately blank. They are already inside the drop-down, "
                    f"the Budget vs Actual table and every dashboard formula — type a category into one and it works immediately.")
note(cat, LAST + 3, "Do not insert or delete rows. Rename what you do not need, or leave it blank.")
note(cat, LAST + 4, "Keep category names unique — the Budget and Dashboard tabs match on the exact text.")
cat.sheet_view.showGridLines = False

# ---------- Transactions ----------
tx = wb.create_sheet("Transactions")
title_block(tx, "Transactions", "Log every dollar in and out. Paste from your bank export or type as you go.", 8)
header_row(tx, 4, ["Date", "Month", "Year", "Category", "Description", "Money In", "Money Out", "Net"])
ROWS = 1200
TXEND = 4 + ROWS
dv = DataValidation(type="list", formula1=f"=Categories!$B${FIRST}:$B${LAST}", allow_blank=True)
tx.add_data_validation(dv)
for i in range(5, 5 + ROWS):
    input_cell(tx, f"A{i}", numfmt='yyyy-mm-dd')
    calc_cell(tx, f"B{i}", f'=IF($A{i}="","",TEXT($A{i},"mmm"))')
    calc_cell(tx, f"C{i}", f'=IF($A{i}="","",YEAR($A{i}))', '0')
    input_cell(tx, f"D{i}")
    input_cell(tx, f"E{i}")
    input_cell(tx, f"F{i}", numfmt='#,##0.00')
    input_cell(tx, f"G{i}", numfmt='#,##0.00')
    calc_cell(tx, f"H{i}", f'=IF(AND($F{i}="",$G{i}=""),"",N($F{i})-N($G{i}))', numfmt='#,##0.00')
    dv.add(tx[f"D{i}"])
widths(tx, {"A": 13, "B": 9, "C": 8, "D": 34, "E": 38, "F": 13, "G": 13, "H": 13})
tx.freeze_panes = "A5"
tx.sheet_view.showGridLines = False
note(tx, 5 + ROWS + 1, "Year is filled in for you from the date. It is what keeps one year's totals separate from the next.")

TXCAT = f"Transactions!$D$5:$D${TXEND}"
TXYR = f"Transactions!$C$5:$C${TXEND}"
TXMON = f"Transactions!$B$5:$B${TXEND}"
TXIN = f"Transactions!$F$5:$F${TXEND}"
TXOUT = f"Transactions!$G$5:$G${TXEND}"

# ---------- Budget vs Actual ----------
bud = wb.create_sheet("Budget vs Actual")
title_block(bud, "Budget vs Actual", "Type your monthly plan. Actuals pull from the Transactions tab for the year you choose.", 16)
bud.cell(row=3, column=1, value="Budget year →").font = Font(bold=True, color=MID)
input_cell(bud, "B3", 2026, '0')
bud.cell(row=3, column=3, value="Only transactions dated in this year are counted. Change it and every actual updates.").font = Font(size=9, italic=True, color=GREY)
YEAR = "'Budget vs Actual'!$B$3"
header_row(bud, 4, ["Category"] + MONTHS + ["Budget Total", "Actual Total", "Variance"])
for k in range(SLOTS):
    r = FIRST + k
    bud.cell(row=r, column=1, value=f"=Categories!B{r}").border = BOX
    for m in range(12):
        input_cell(bud, f"{get_column_letter(2+m)}{r}", 0, numfmt='#,##0')
    calc_cell(bud, f"N{r}", f"=SUM(B{r}:M{r})", '#,##0', bold=True)
    calc_cell(bud, f"O{r}",
              f'=IF($A{r}="","",SUMIFS({TXIN},{TXCAT},$A{r},{TXYR},$B$3)'
              f'+SUMIFS({TXOUT},{TXCAT},$A{r},{TXYR},$B$3))', '#,##0', bold=True)
    calc_cell(bud, f"P{r}", f'=IF($A{r}="","",N{r}-O{r})', '#,##0')
TOT = LAST + 1
bud.cell(row=TOT, column=1, value="TOTAL").font = Font(bold=True)
for col in range(2, 17):
    L = get_column_letter(col)
    calc_cell(bud, f"{L}{TOT}", f"=SUM({L}{FIRST}:{L}{TOT-1})", '#,##0', bold=True)
widths(bud, dict({"A": 34, "N": 13, "O": 13, "P": 12}, **{get_column_letter(2+i): 10 for i in range(12)}))
bud.freeze_panes = "B5"
bud.sheet_view.showGridLines = False

# ---------- Dashboard ----------
d = wb.create_sheet("Dashboard")
title_block(d, "Dashboard", "Your year at a glance — updates itself as you log transactions.", 6)
d.cell(row=3, column=1, value="Showing budget year").font = Font(bold=True, color=MID)
calc_cell(d, "B3", f"={YEAR}", '0', bold=True)
section(d, 4, "THE HEADLINE NUMBERS", 6)
saved = "+".join(f'SUMIFS({TXOUT},{TXCAT},"{c}",{TXYR},$B$3)'
                 for c in ["TFSA contribution", "RRSP contribution", "RESP contribution", "Emergency fund"])
rows = [
    ("Total money in", f'=SUMIFS({TXIN},{TXYR},$B$3)'),
    ("Total money out", f'=SUMIFS({TXOUT},{TXYR},$B$3)'),
    ("Net cash flow", "=B6-B7"),
    ("Total saved / invested", "=" + saved),
    ("Savings rate", '=IF(B6=0,"—",B9/B6)'),
    ("Budgeted total", f"='Budget vs Actual'!N{TOT}"),
    ("Actual total", f"='Budget vs Actual'!O{TOT}"),
]
r = 6
for label, f in rows:
    d.cell(row=r, column=1, value=label).font = Font(bold=True)
    calc_cell(d, f"B{r}", f, '0.0%' if label == "Savings rate" else '$#,##0.00', bold=True)
    r += 1
section(d, r + 1, "MONTH BY MONTH", 6)
mr = r + 2
header_row(d, mr, ["Month", "In", "Out", "Net"])
for i, m in enumerate(MONTHS):
    rr = mr + 1 + i
    d.cell(row=rr, column=1, value=m).border = BOX
    calc_cell(d, f"B{rr}", f'=SUMIFS({TXIN},{TXMON},$A{rr},{TXYR},$B$3)', '#,##0.00')
    calc_cell(d, f"C{rr}", f'=SUMIFS({TXOUT},{TXMON},$A{rr},{TXYR},$B$3)', '#,##0.00')
    calc_cell(d, f"D{rr}", f"=B{rr}-C{rr}", '#,##0.00', bold=True)
ch = BarChart(); ch.type = "col"; ch.title = "Money in vs money out by month"; ch.height = 8; ch.width = 18
ch.add_data(Reference(d, min_col=2, max_col=3, min_row=mr, max_row=mr + 12), titles_from_data=True)
ch.set_categories(Reference(d, min_col=1, min_row=mr + 1, max_row=mr + 12))
d.add_chart(ch, "F6")
widths(d, {"A": 26, "B": 16, "C": 16, "D": 16})
d.sheet_view.showGridLines = False

start_here(wb, "Canadian Budget & Cash-Flow Tracker",
    "Built by a Canadian personal banker. Track every dollar, see your real savings rate.",
    ["Open the Categories tab and rename anything that does not fit your life. Spare blank rows are waiting at the bottom.",
     "Go to Budget vs Actual, set the budget year in the yellow cell at the top, and type what you PLAN to spend each month.",
     "Log spending on the Transactions tab — date, category, money in or money out.",
     "The Dashboard and the Actual columns update themselves. Nothing else to do.",
     "Keep several years in the same file: change the budget year and every total follows it."],
    ["Start Here — this page.",
     f"Categories — {len(CATS)} Canadian budget categories plus {SPARES} spare slots, all pre-wired.",
     "Transactions — 1,200-row logbook with a drop-down category picker.",
     "Budget vs Actual — 12-month plan against real spending, filtered to one budget year.",
     "Dashboard — income, spending, net cash flow, savings rate and a monthly chart."])
finish(wb, os.path.join(OUT, "01-Canadian-Budget-and-Cash-Flow-Tracker.xlsx"))
