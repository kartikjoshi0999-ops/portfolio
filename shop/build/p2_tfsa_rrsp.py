import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from kjstyle import *

OUT = os.path.join(os.path.dirname(__file__), "..", "dist")
TFSA = [(2009,5000),(2010,5000),(2011,5000),(2012,5000),(2013,5500),(2014,5500),(2015,10000),
        (2016,5500),(2017,5500),(2018,5500),(2019,6000),(2020,6000),(2021,6000),(2022,6000),
        (2023,6500),(2024,7000),(2025,7000),(2026,7000)]
RRSP = [(2019,26500),(2020,27230),(2021,27830),(2022,29210),(2023,30780),(2024,31560),
        (2025,32490),(2026,33810),(2027,35130)]

wb = Workbook(); wb.remove(wb.active)

# ---------- Inputs ----------
i = wb.create_sheet("Your Numbers")
title_block(i, "Your Numbers", "Fill in the yellow cells. Everything else calculates.", 5)
section(i, 4, "ABOUT YOU", 5)
about = [
    ("Your name", "", None),
    ("Current age", 30, '0'),
    ("Year you turned 18", 2013, '0'),
    ("Year you became a Canadian resident", 2013, '0'),
    ("Planning year", 2026, '0'),
    ("Province / territory", "Ontario", None),
]
r = 5
for lab, val, fmt in about:
    i.cell(row=r, column=1, value=lab).font = Font(bold=True)
    input_cell(i, f"B{r}", val, fmt); r += 1
note(i, r, "TFSA room starts in the LATER of the year you turned 18 and the year you became a Canadian resident.")
note(i, r + 1, "If you were born in Canada, put the same year in both boxes.")
r += 3
section(i, r, "INCOME & TAX", 5); r += 1
INC_ROW = r
tax = [
    ("Earned income LAST year (employment + net self-employment)", 78000, '$#,##0'),
    ("Marginal tax rate NOW (combined fed + prov)", 0.2965, '0.00%'),
    ("Expected marginal tax rate IN RETIREMENT", 0.2005, '0.00%'),
    ("Pension adjustment last year (box 52 of your T4)", 0, '$#,##0'),
]
for lab, val, fmt in tax:
    i.cell(row=r, column=1, value=lab).font = Font(bold=True)
    input_cell(i, f"B{r}", val, fmt); r += 1
note(i, r, "RRSP room is earned on LAST year's income and becomes available THIS year — that is why last year's figure is what belongs here.")
note(i, r + 1, "Marginal rate = the tax you pay on your NEXT dollar of income, not your average rate.")
r += 3
section(i, r, "WHAT YOU HAVE ALREADY", 5); r += 1
HAVE_ROW = r
have = [
    ("Total TFSA contributions made to date", 21000, '$#,##0'),
    ("Total TFSA withdrawals in PREVIOUS years", 0, '$#,##0'),
    ("RRSP deduction limit for this year (from your CRA notice)", 44000, '$#,##0'),
    ("RRSP contributions already made this year", 3000, '$#,##0'),
]
for lab, val, fmt in have:
    i.cell(row=r, column=1, value=lab).font = Font(bold=True)
    input_cell(i, f"B{r}", val, fmt); r += 1
note(i, r, "Your CRA My Account notice of assessment is the authoritative source. The RRSP deduction limit it shows ALREADY includes your carry-forward — do not add anything to it.")
r += 3
section(i, r, "PLANNING ASSUMPTIONS", 5); r += 1
ASS_ROW = r
ass = [
    ("Amount you can save per year", 12000, '$#,##0'),
    ("Expected annual investment return", 0.06, '0.00%'),
    ("Years until you draw the money", 25, '0'),
]
for lab, val, fmt in ass:
    i.cell(row=r, column=1, value=lab).font = Font(bold=True)
    input_cell(i, f"B{r}", val, fmt); r += 1
widths(i, {"A": 56, "B": 18})
i.sheet_view.showGridLines = False

AGE, T18, TRES, PY = ["'Your Numbers'!$B$%d" % n for n in (6, 7, 8, 9)]
INC, MTR_NOW, MTR_RET, PA = [f"'Your Numbers'!$B${INC_ROW+k}" for k in range(4)]
TFSA_MADE, TFSA_WD, RRSP_LIMIT, RRSP_MADE = [f"'Your Numbers'!$B${HAVE_ROW+k}" for k in range(4)]
SAVE, RET, YRS = [f"'Your Numbers'!$B${ASS_ROW+k}" for k in range(3)]

# ---------- TFSA ----------
t = wb.create_sheet("TFSA Room")
title_block(t, "TFSA Room", "Year-by-year room, counted from the later of age 18 and Canadian residency.", 4)
header_row(t, 4, ["Year", "Annual TFSA limit", "Counts for you?", "Room earned"])
r = 5
for yr, lim in TFSA:
    t.cell(row=r, column=1, value=yr).border = BOX
    t.cell(row=r, column=2, value=lim).border = BOX; t[f"B{r}"].number_format = '$#,##0'
    calc_cell(t, f"C{r}", f'=IF(AND(A{r}>=MAX({T18},{TRES}),A{r}<={PY}),"Yes","No")')
    calc_cell(t, f"D{r}", f'=IF(C{r}="Yes",B{r},0)', '$#,##0')
    r += 1
END = r - 1
r += 1
t.cell(row=r, column=1, value="Total room earned").font = Font(bold=True)
calc_cell(t, f"B{r}", f"=SUM(D5:D{END})", '$#,##0', bold=True); TOTROOM = r; r += 1
t.cell(row=r, column=1, value="Less: contributions made").font = Font(bold=True)
calc_cell(t, f"B{r}", f"={TFSA_MADE}", '$#,##0'); r += 1
t.cell(row=r, column=1, value="Add back: prior-year withdrawals").font = Font(bold=True)
calc_cell(t, f"B{r}", f"={TFSA_WD}", '$#,##0'); r += 1
t.cell(row=r, column=1, value="ROOM AVAILABLE TODAY").font = Font(bold=True, size=12, color=MID)
calc_cell(t, f"B{r}", f"=MAX(0,B{TOTROOM}-B{TOTROOM+1}+B{TOTROOM+2})", '$#,##0', bold=True)
TFSA_AVAIL = f"'TFSA Room'!$B${r}"; r += 2
note(t, r, "2026 limit shown as $7,000 — confirm on canada.ca before you rely on it; the limit is indexed each year.")
note(t, r+1, "Money withdrawn from a TFSA comes back as room on January 1 of the FOLLOWING year, never the same year.")
note(t, r+2, "Over-contributing costs 1% of the excess per month. Check CRA My Account if you are close to the line.")
note(t, r+3, "Newcomers: room does NOT build for years you were not a Canadian resident, even if you were over 18.")
widths(t, {"A": 12, "B": 22, "C": 18, "D": 16})
t.sheet_view.showGridLines = False

# ---------- RRSP ----------
q = wb.create_sheet("RRSP Room")
title_block(q, "RRSP Room", "What you can contribute today, and what next year's room will be.", 4)
header_row(q, 4, ["Year", "Annual RRSP dollar limit"])
r = 5
for yr, lim in RRSP:
    q.cell(row=r, column=1, value=yr).border = BOX
    q.cell(row=r, column=2, value=lim).border = BOX; q[f"B{r}"].number_format = '$#,##0'
    r += 1
LIMEND = r - 1
r += 1
section(q, r, "ROOM AVAILABLE TO YOU TODAY", 4); r += 1
TODAY = r
q.cell(row=r, column=1, value="RRSP deduction limit for this year (from your CRA notice)").font = Font(bold=True)
calc_cell(q, f"B{r}", f"={RRSP_LIMIT}", '$#,##0'); r += 1
q.cell(row=r, column=1, value="Less: contributions already made this year").font = Font(bold=True)
calc_cell(q, f"B{r}", f"={RRSP_MADE}", '$#,##0'); r += 1
q.cell(row=r, column=1, value="ROOM AVAILABLE TODAY").font = Font(bold=True, size=12, color=MID)
calc_cell(q, f"B{r}", f"=MAX(0,B{TODAY}-B{TODAY+1})", '$#,##0', bold=True)
RRSP_AVAIL = f"'RRSP Room'!$B${r}"; r += 1
q.cell(row=r, column=1, value="Tax refund if you contribute it all").font = Font(bold=True)
calc_cell(q, f"B{r}", f"=B{TODAY+2}*{MTR_NOW}", '$#,##0', bold=True); r += 2
note(q, r, "The deduction limit on your notice of assessment ALREADY includes every unused dollar carried forward. Adding this year's new room to it would double-count.")
r += 2
section(q, r, "ROOM YOU WILL EARN FOR NEXT YEAR", 4); r += 1
NEXT = r
q.cell(row=r, column=1, value="18% of last year's earned income").font = Font(bold=True)
calc_cell(q, f"B{r}", f"=0.18*{INC}", '$#,##0'); r += 1
q.cell(row=r, column=1, value="Dollar limit for next year").font = Font(bold=True)
calc_cell(q, f"B{r}", f'=IFERROR(VLOOKUP({PY}+1,A5:B{LIMEND},2,FALSE),MAX(B5:B{LIMEND}))', '$#,##0'); r += 1
q.cell(row=r, column=1, value="New room earned (lesser of the two)").font = Font(bold=True)
calc_cell(q, f"B{r}", f"=MIN(B{NEXT},B{NEXT+1})", '$#,##0'); r += 1
q.cell(row=r, column=1, value="Less: pension adjustment").font = Font(bold=True)
calc_cell(q, f"B{r}", f"={PA}", '$#,##0'); r += 1
q.cell(row=r, column=1, value="ROOM ADDED FOR NEXT YEAR").font = Font(bold=True, size=12, color=MID)
calc_cell(q, f"B{r}", f"=MAX(0,B{NEXT+2}-B{NEXT+3})", '$#,##0', bold=True); r += 2
note(q, r, "Contributions earn room on the PREVIOUS year's income: 18% of what you earned last year is what became available this year.")
note(q, r+1, "Dollar limits are as published by CRA; later years are indexed — verify before filing.")
note(q, r+2, "You have a $2,000 lifetime over-contribution cushion, but it is not deductible. Do not lean on it.")
widths(q, {"A": 54, "B": 20})
q.sheet_view.showGridLines = False

# ---------- Compare ----------
c = wb.create_sheet("TFSA vs RRSP")
title_block(c, "TFSA vs RRSP", "Same dollars, both accounts, side by side after tax — capped at the room you actually have.", 6)
section(c, 4, "IF YOU PUT THIS YEAR'S SAVINGS IN ONE ACCOUNT", 6)
header_row(c, 5, ["", "TFSA route", "RRSP route", "What this line means"])
rows = [
    ("Amount you want to save", f"={SAVE}", f"={SAVE}", "What you told the Your Numbers tab you can put away."),
    ("Room available in this account", f"={TFSA_AVAIL}", f"={RRSP_AVAIL}", "Pulled from the two room tabs. This is the ceiling."),
    ("Amount actually contributed", "=MIN(B6,B7)", "=MIN(C6,C7)", "Capped at your room — the planner will not model an over-contribution."),
    ("Immediate tax refund", "0", f"=C8*{MTR_NOW}", "RRSP contributions are deductible; TFSA contributions are not."),
    ("Refund reinvested at the same return", "0", f"=C9*(1+{RET})^{YRS}", "Only counts if you actually invest the refund."),
    ("Value at the end of the horizon", f"=B8*(1+{RET})^{YRS}", f"=C8*(1+{RET})^{YRS}", "Growth before any tax on withdrawal."),
    ("Tax on withdrawal", "0", f"=-C11*{MTR_RET}", "RRSP/RRIF withdrawals are fully taxable as income."),
    ("Refund pot after tax (non-registered, simplified)", "0", f"=C10*(1-{MTR_RET}*0.5)", "Assumes the refund grows in a taxable account."),
    ("AFTER-TAX VALUE", "=B11", "=C11+C12+C13", "The number that actually matters."),
]
r = 6
for lab, bf, cf, why in rows:
    bold = lab == "AFTER-TAX VALUE"
    c.cell(row=r, column=1, value=lab).font = Font(bold=bold)
    calc_cell(c, f"B{r}", bf, '$#,##0', bold=bold)
    calc_cell(c, f"C{r}", cf, '$#,##0', bold=bold)
    c.cell(row=r, column=4, value=why).font = Font(size=9, italic=True, color=GREY)
    r += 1
r += 1
c.cell(row=r, column=1, value="WHICH ONE WINS FOR YOU").font = Font(bold=True, size=12, color=MID)
r += 1
verdict = (
    f'=IF(AND({TFSA_AVAIL}<=0,{RRSP_AVAIL}<=0),'
    f'"You have no room left in either account this year. Anything more goes to a non-registered account, or waits for January.",'
    f'IF({RRSP_AVAIL}<=0,"TFSA — you have no RRSP room left this year.",'
    f'IF({TFSA_AVAIL}<=0,"RRSP — you have no TFSA room left this year.",'
    f'IF({MTR_NOW}>{MTR_RET},"RRSP first — you deduct at a higher rate than you will pay on withdrawal."'
    f'&IF({RRSP_AVAIL}<{SAVE}," Your RRSP room runs out before your savings do, so put the remainder in your TFSA.",""),'
    f'IF({MTR_NOW}<{MTR_RET},"TFSA first — you would deduct at a low rate now and pay tax at a higher one later."'
    f'&IF({TFSA_AVAIL}<{SAVE}," Your TFSA room runs out before your savings do, so put the remainder in your RRSP.",""),'
    f'"Close call — your rate now equals your expected rate later, so use the TFSA for its flexibility.")))))'
)
calc_cell(c, f"A{r}", verdict)
c.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c[f"A{r}"].font = Font(bold=True, size=11, color=MID)
c[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="center")
c.row_dimensions[r].height = 38
r += 2
section(c, r, "YOUR ROOM, PULLED FROM THE OTHER TABS", 6); r += 1
c.cell(row=r, column=1, value="TFSA room available").font = Font(bold=True)
calc_cell(c, f"B{r}", f"={TFSA_AVAIL}", '$#,##0', bold=True); r += 1
c.cell(row=r, column=1, value="RRSP room available").font = Font(bold=True)
calc_cell(c, f"B{r}", f"={RRSP_AVAIL}", '$#,##0', bold=True); r += 1
c.cell(row=r, column=1, value="Total room available").font = Font(bold=True)
calc_cell(c, f"B{r}", f"=B{r-2}+B{r-1}", '$#,##0', bold=True); r += 1
c.cell(row=r, column=1, value="Savings with nowhere registered to go").font = Font(bold=True)
calc_cell(c, f"B{r}", f"=MAX(0,{SAVE}-B{r-1})", '$#,##0', bold=True)
widths(c, {"A": 46, "B": 18, "C": 18, "D": 52})
c.sheet_view.showGridLines = False

# ---------- Projection ----------
p = wb.create_sheet("Projection")
title_block(p, "Projection", "What steady contributions look like over time.", 5)
header_row(p, 4, ["Year", "Contributed this year", "Cumulative contributed", "Growth", "Balance"])
for k in range(30):
    r = 5 + k
    p.cell(row=r, column=1, value=k + 1).border = BOX
    calc_cell(p, f"B{r}", f'=IF(A{r}<={YRS},{SAVE},0)', '$#,##0')
    calc_cell(p, f"C{r}", (f"=B{r}" if k == 0 else f"=C{r-1}+B{r}"), '$#,##0')
    prev = "0" if k == 0 else f"E{r-1}"
    calc_cell(p, f"D{r}", f"=({prev}+B{r})*{RET}", '$#,##0')
    calc_cell(p, f"E{r}", f"={prev}+B{r}+D{r}", '$#,##0', bold=True)
ch = LineChart(); ch.title = "Projected balance"; ch.height = 9; ch.width = 20
ch.add_data(Reference(p, min_col=5, min_row=4, max_row=34), titles_from_data=True)
ch.set_categories(Reference(p, min_col=1, min_row=5, max_row=34))
p.add_chart(ch, "G5")
widths(p, {"A": 8, "B": 20, "C": 22, "D": 14, "E": 16})
p.sheet_view.showGridLines = False

start_here(wb, "TFSA vs RRSP Contribution Planner (Canada)",
    "Know your real room, and know which account to fill first.",
    ["Fill in the yellow cells on the Your Numbers tab — that is the only typing you do.",
     "Open TFSA Room: room builds from the later of the year you turned 18 and the year you became a resident.",
     "Open RRSP Room: today's room comes straight off your CRA notice; next year's is 18% of last year's income.",
     "Open TFSA vs RRSP for the after-tax answer — capped at the room you actually have.",
     "Use Projection to see what steady contributions turn into over your time horizon."],
    ["Your Numbers — every input in one place.",
     "TFSA Room — 2009-2026 limits, counted from age 18 AND Canadian residency.",
     "RRSP Room — today's room from your CRA notice, and next year's earned room.",
     "TFSA vs RRSP — after-tax comparison capped at your real room, with a plain-English verdict.",
     "Projection — 30-year contribution and growth schedule with a chart."])
finish(wb, os.path.join(OUT, "02-TFSA-vs-RRSP-Contribution-Planner.xlsx"))
