import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, LineChart, Reference
from kjstyle import *

OUT = os.path.join(os.path.dirname(__file__), "..", "dist")
H1, H2 = 5, 64          # holdings rows
wb = Workbook(); wb.remove(wb.active)

# ---------- Holdings ----------
h = wb.create_sheet("Holdings")
title_block(h, "Holdings", "One row per position. Update the price column and everything else follows.", 12)
header_row(h, 4, ["Ticker", "Name", "Account", "Asset class", "Currency", "Units",
                  "Avg cost / unit", "Book value", "Current price", "Market value",
                  "Gain / loss $", "Gain / loss %", "% of portfolio"])
sample = [
    ("VFV.TO", "Vanguard S&P 500 Index ETF", "TFSA", "Equity — US", "CAD", 120, 118.40, 141.20),
    ("XIC.TO", "iShares Core S&P/TSX Capped ETF", "TFSA", "Equity — Canada", "CAD", 200, 34.10, 39.85),
    ("XAW.TO", "iShares Core MSCI All Country ex-Canada", "RRSP", "Equity — Global", "CAD", 150, 32.75, 38.40),
    ("ZAG.TO", "BMO Aggregate Bond Index ETF", "RRSP", "Fixed income", "CAD", 180, 15.20, 14.60),
    ("CASH.TO", "High Interest Savings ETF", "Non-registered", "Cash & equivalents", "CAD", 90, 50.05, 50.12),
    ("RY.TO", "Royal Bank of Canada", "Non-registered", "Equity — Canada", "CAD", 40, 121.50, 168.30),
]
acc = DataValidation(type="list", formula1='"TFSA,RRSP,FHSA,RESP,Non-registered,LIRA,Other"', allow_blank=True)
cls = DataValidation(type="list", formula1='"Equity — Canada,Equity — US,Equity — Global,Fixed income,Cash & equivalents,Real assets,Crypto,Other"', allow_blank=True)
h.add_data_validation(acc); h.add_data_validation(cls)
for r in range(H1, H2 + 1):
    k = r - H1
    s = sample[k] if k < len(sample) else ("",) * 8
    for col, val in zip("ABCDE", s[:5]):
        input_cell(h, f"{col}{r}", val or None)
    input_cell(h, f"F{r}", s[5] or None, '#,##0.0000')
    input_cell(h, f"G{r}", s[6] or None, '#,##0.0000')
    calc_cell(h, f"H{r}", f'=IF($A{r}="","",N(F{r})*N(G{r}))', '$#,##0.00')
    input_cell(h, f"I{r}", s[7] or None, '#,##0.0000')
    calc_cell(h, f"J{r}", f'=IF($A{r}="","",N(F{r})*N(I{r}))', '$#,##0.00')
    calc_cell(h, f"K{r}", f'=IF($A{r}="","",N(J{r})-N(H{r}))', '$#,##0.00')
    calc_cell(h, f"L{r}", f'=IF(OR($A{r}="",N(H{r})=0),"",N(K{r})/N(H{r}))', '0.0%')
    calc_cell(h, f"M{r}", f'=IF(OR($A{r}="",$J${H2+1}=0),"",N(J{r})/$J${H2+1})', '0.0%')
    acc.add(h[f"C{r}"]); cls.add(h[f"D{r}"])
T = H2 + 1
h.cell(row=T, column=1, value="TOTAL").font = Font(bold=True)
for col in "HJK":
    calc_cell(h, f"{col}{T}", f"=SUM({col}{H1}:{col}{H2})", '$#,##0.00', bold=True)
calc_cell(h, f"L{T}", f'=IF(H{T}=0,"",K{T}/H{T})', '0.0%', bold=True)
widths(h, {"A": 11, "B": 36, "C": 15, "D": 20, "E": 10, "F": 11, "G": 14,
           "H": 14, "I": 13, "J": 14, "K": 14, "L": 12, "M": 13})
h.freeze_panes = "A5"; h.sheet_view.showGridLines = False

# ---------- Contributions ----------
co = wb.create_sheet("Contributions")
title_block(co, "Contributions", "Every deposit and withdrawal, by account. Needed for real return.", 5)
header_row(co, 4, ["Date", "Account", "Type", "Amount", "Note"])
tdv = DataValidation(type="list", formula1='"Contribution,Withdrawal,Dividend received,Fee"', allow_blank=True)
co.add_data_validation(tdv)
for r in range(5, 305):
    input_cell(co, f"A{r}", numfmt='yyyy-mm-dd'); input_cell(co, f"B{r}")
    input_cell(co, f"C{r}"); input_cell(co, f"D{r}", numfmt='$#,##0.00'); input_cell(co, f"E{r}")
    tdv.add(co[f"C{r}"])
widths(co, {"A": 13, "B": 18, "C": 20, "D": 15, "E": 44})
co.freeze_panes = "A5"; co.sheet_view.showGridLines = False

# ---------- Dividends ----------
dv = wb.create_sheet("Dividend Log")
title_block(dv, "Dividend Log", "Track income received so you can see your real yield.", 6)
header_row(dv, 4, ["Date", "Ticker", "Account", "Amount", "Withholding tax", "Net received"])
for r in range(5, 205):
    input_cell(dv, f"A{r}", numfmt='yyyy-mm-dd'); input_cell(dv, f"B{r}"); input_cell(dv, f"C{r}")
    input_cell(dv, f"D{r}", numfmt='$#,##0.00'); input_cell(dv, f"E{r}", numfmt='$#,##0.00')
    calc_cell(dv, f"F{r}", f'=IF($A{r}="","",N(D{r})-N(E{r}))', '$#,##0.00')
dv.cell(row=206, column=1, value="TOTAL").font = Font(bold=True)
for col in "DEF":
    calc_cell(dv, f"{col}206", f"=SUM({col}5:{col}205)", '$#,##0.00', bold=True)
widths(dv, {"A": 13, "B": 12, "C": 18, "D": 15, "E": 16, "F": 15})
dv.freeze_panes = "A5"; dv.sheet_view.showGridLines = False

# ---------- Dashboard ----------
d = wb.create_sheet("Dashboard")
title_block(d, "Dashboard", "Where you stand, what it cost, and how it is split.", 6)
section(d, 4, "PORTFOLIO AT A GLANCE", 6)
rows = [
    ("Market value", f"=Holdings!J{T}", '$#,##0.00'),
    ("Book value (what you paid)", f"=Holdings!H{T}", '$#,##0.00'),
    ("Unrealised gain / loss", f"=Holdings!K{T}", '$#,##0.00'),
    ("Return on book", f'=IF(B7=0,"—",B8/B7)', '0.0%'),
    ("Dividends received (net)", "='Dividend Log'!F206", '$#,##0.00'),
    ("Net contributions", '=SUMIF(Contributions!$C$5:$C$304,"Contribution",Contributions!$D$5:$D$304)'
                          '-SUMIF(Contributions!$C$5:$C$304,"Withdrawal",Contributions!$D$5:$D$304)', '$#,##0.00'),
    ("Total gain incl. dividends", "=B8+B10", '$#,##0.00'),
    ("Number of positions", f'=COUNTA(Holdings!A{H1}:A{H2})', '0'),
]
r = 6
for lab, f, fmt in rows:
    d.cell(row=r, column=1, value=lab).font = Font(bold=True)
    calc_cell(d, f"B{r}", f, fmt, bold=True); r += 1
r += 1
section(d, r, "ASSET ALLOCATION", 6); r += 1
AR = r
header_row(d, r, ["Asset class", "Market value", "% of total", "Target %", "Drift"]); r += 1
classes = ["Equity — Canada", "Equity — US", "Equity — Global", "Fixed income",
           "Cash & equivalents", "Real assets", "Crypto", "Other"]
targets = [0.20, 0.30, 0.20, 0.20, 0.05, 0.03, 0.02, 0.00]
for cname, tg in zip(classes, targets):
    d.cell(row=r, column=1, value=cname).border = BOX
    calc_cell(d, f"B{r}", f'=SUMIF(Holdings!$D${H1}:$D${H2},$A{r},Holdings!$J${H1}:$J${H2})', '$#,##0.00')
    calc_cell(d, f"C{r}", f'=IF($B$6=0,0,B{r}/$B$6)', '0.0%')
    input_cell(d, f"D{r}", tg, '0.0%')
    calc_cell(d, f"E{r}", f"=C{r}-D{r}", '0.0%')
    r += 1
d.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
for col, fmt in (("B", '$#,##0.00'), ("C", '0.0%'), ("D", '0.0%')):
    calc_cell(d, f"{col}{r}", f"=SUM({col}{AR+1}:{col}{r-1})", fmt, bold=True)
ALLOC_END = r - 1
r += 2
section(d, r, "BY ACCOUNT", 6); r += 1
ACC_R = r
header_row(d, r, ["Account", "Market value", "% of total"]); r += 1
for a in ["TFSA", "RRSP", "FHSA", "RESP", "Non-registered", "LIRA", "Other"]:
    d.cell(row=r, column=1, value=a).border = BOX
    calc_cell(d, f"B{r}", f'=SUMIF(Holdings!$C${H1}:$C${H2},$A{r},Holdings!$J${H1}:$J${H2})', '$#,##0.00')
    calc_cell(d, f"C{r}", f'=IF($B$6=0,0,B{r}/$B$6)', '0.0%')
    r += 1
pie = PieChart(); pie.title = "Asset allocation"; pie.height = 9; pie.width = 12
pie.add_data(Reference(d, min_col=2, min_row=AR, max_row=ALLOC_END), titles_from_data=True)
pie.set_categories(Reference(d, min_col=1, min_row=AR + 1, max_row=ALLOC_END))
d.add_chart(pie, "H6")
pie2 = PieChart(); pie2.title = "By account"; pie2.height = 9; pie2.width = 12
pie2.add_data(Reference(d, min_col=2, min_row=ACC_R, max_row=ACC_R + 7), titles_from_data=True)
pie2.set_categories(Reference(d, min_col=1, min_row=ACC_R + 1, max_row=ACC_R + 7))
d.add_chart(pie2, "H26")
widths(d, {"A": 30, "B": 18, "C": 13, "D": 12, "E": 12})
d.sheet_view.showGridLines = False

start_here(wb, "Investment Portfolio Tracker (Canada)",
    "TFSA, RRSP, FHSA and non-registered in one place, with drift against your targets.",
    ["Enter your positions on the Holdings tab — units, average cost and today's price.",
     "Pick the account and asset class from the drop-downs; the Dashboard groups on them.",
     "Log deposits and withdrawals on Contributions so you can see real money in versus growth.",
     "Record dividends on the Dividend Log — including withholding tax on US names.",
     "Set your target weights on the Dashboard. The drift column tells you what to rebalance."],
    ["Start Here — this page.",
     "Holdings — 60 positions with book value, market value and gain/loss.",
     "Contributions — 300-row deposit and withdrawal register.",
     "Dividend Log — 200-row income tracker with withholding tax.",
     "Dashboard — totals, allocation vs target with drift, split by account, two charts."])
finish(wb, os.path.join(OUT, "04-Investment-Portfolio-Tracker.xlsx"))
