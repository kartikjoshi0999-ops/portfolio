import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import LineChart, BarChart, Reference
from kjstyle import *

OUT = os.path.join(os.path.dirname(__file__), "..", "dist")
wb = Workbook(); wb.remove(wb.active)
COLS = [get_column_letter(2 + i) for i in range(12)]   # B..M
TOTC = "N"

def month_headers(ws, row, first_label="Line item"):
    header_row(ws, row, [first_label] + [f"=TEXT(EDATE(Assumptions!$B$5,{i}),\"mmm yy\")" for i in range(12)] + ["Total"])

# ---------- Assumptions ----------
a = wb.create_sheet("Assumptions")
title_block(a, "Assumptions", "Change these and the whole forecast moves. Nothing else needs touching.", 6)
section(a, 4, "BUSINESS SETUP", 6)
setup = [("Forecast start month", "2026-01-01", 'mmm yyyy'),
         ("Business name", "My Business Inc.", None),
         ("Opening cash balance", 25000, '$#,##0'),
         ("Sales collected in the month of sale", 0.60, '0%'),
         ("Sales collected 30 days later", 0.35, '0%'),
         ("Sales never collected (bad debt)", 0.05, '0%'),
         ("Supplier terms — paid in month", 0.70, '0%'),
         ("Supplier terms — paid 30 days later", 0.30, '0%'),
         ("Cost of goods sold as % of revenue", 0.42, '0%'),
         ("Sales tax collected on revenue (HST)", 0.13, '0%'),
         ("Sales tax remitted quarterly?", "Yes", None),
         ("Corporate tax rate on profit", 0.12, '0%'),
         ("Monthly revenue growth", 0.03, '0.0%'),
         ("Month 1 revenue", 48000, '$#,##0')]
r = 5
for lab, val, fmt in setup:
    a.cell(row=r, column=1, value=lab).font = Font(bold=True)
    input_cell(a, f"B{r}", val, fmt); r += 1
note(a, r, "Collection percentages should add to 100%. Supplier terms should add to 100%.")
calc_cell(a, f"B{r+1}", '=IF(ROUND(B8+B9+B10,4)=1,"Collections OK","⚠ Collections do not add to 100%")')
calc_cell(a, f"B{r+2}", '=IF(ROUND(B11+B12,4)=1,"Supplier terms OK","⚠ Supplier terms do not add to 100%")')
widths(a, {"A": 42, "B": 20})
a.sheet_view.showGridLines = False
GROWTH, REV1, COGSPC, HST, TAXR = "Assumptions!$B$17", "Assumptions!$B$18", "Assumptions!$B$13", "Assumptions!$B$14", "Assumptions!$B$16"
CIM, C30, BAD = "Assumptions!$B$8", "Assumptions!$B$9", "Assumptions!$B$10"
SIM, S30 = "Assumptions!$B$11", "Assumptions!$B$12"
OPEN = "Assumptions!$B$7"

# ---------- Revenue ----------
rv = wb.create_sheet("Revenue")
title_block(rv, "Revenue", "Build the top line by stream. Growth applies to the base case.", 14)
month_headers(rv, 4, "Revenue stream")
streams = ["Core product / service", "Recurring / retainer", "Add-ons & upsells", "Other income"]
splits = [0.62, 0.22, 0.11, 0.05]
r = 5
for name, sp in zip(streams, splits):
    input_cell(rv, f"A{r}", name)
    for i, c in enumerate(COLS):
        if i == 0:
            calc_cell(rv, f"{c}{r}", f"={REV1}*{sp}", '$#,##0')
        else:
            calc_cell(rv, f"{c}{r}", f"={COLS[i-1]}{r}*(1+{GROWTH})", '$#,##0')
    calc_cell(rv, f"{TOTC}{r}", f"=SUM(B{r}:M{r})", '$#,##0', bold=True)
    r += 1
REV_T = r
rv.cell(row=r, column=1, value="TOTAL REVENUE").font = Font(bold=True)
for c in COLS + [TOTC]:
    calc_cell(rv, f"{c}{r}", f"=SUM({c}5:{c}{r-1})", '$#,##0', bold=True)
r += 2
section(rv, r, "UNIT ECONOMICS (OPTIONAL)", 14); r += 1
for lab, formula, fmt in [("Customers served", "=40*(1+%s)^{i}" % GROWTH.replace('Assumptions!$B$17','Assumptions!$B$17'), '#,##0'),
                          ("Average revenue per customer", None, '$#,##0.00')]:
    rv.cell(row=r, column=1, value=lab).font = Font(bold=True)
    for i, c in enumerate(COLS):
        if lab == "Customers served":
            input_cell(rv, f"{c}{r}", 40, '#,##0')
        else:
            calc_cell(rv, f"{c}{r}", f'=IF({c}{r-1}=0,0,{c}{REV_T}/{c}{r-1})', '$#,##0.00')
    r += 1
widths(rv, dict({"A": 30, TOTC: 14}, **{c: 12 for c in COLS}))
rv.sheet_view.showGridLines = False

# ---------- Costs ----------
cs = wb.create_sheet("Costs")
title_block(cs, "Costs", "Direct costs move with revenue. Overheads are what you type.", 14)
month_headers(cs, 4, "Cost line")
r = 5
cs.cell(row=r, column=1, value="Cost of goods sold").font = Font(bold=True)
for c in COLS:
    calc_cell(cs, f"{c}{r}", f"=Revenue!{c}{REV_T}*{COGSPC}", '$#,##0')
calc_cell(cs, f"{TOTC}{r}", f"=SUM(B{r}:M{r})", '$#,##0', bold=True)
COGS_R = r; r += 2
section(cs, r, "OPERATING EXPENSES", 14); r += 1
OPEX_START = r
opex = [("Salaries & wages", 14000), ("Payroll taxes & benefits", 2100), ("Rent", 3200),
        ("Utilities", 480), ("Insurance", 390), ("Software & subscriptions", 620),
        ("Marketing & advertising", 2400), ("Professional fees (legal, accounting)", 750),
        ("Bank & merchant fees", 540), ("Repairs & maintenance", 300),
        ("Travel & vehicle", 650), ("Office supplies", 220),
        ("Training & development", 250), ("Other overhead", 400)]
for name, amt in opex:
    input_cell(cs, f"A{r}", name)
    for c in COLS:
        input_cell(cs, f"{c}{r}", amt, '$#,##0')
    calc_cell(cs, f"{TOTC}{r}", f"=SUM(B{r}:M{r})", '$#,##0', bold=True)
    r += 1
OPEX_T = r
cs.cell(row=r, column=1, value="TOTAL OPERATING EXPENSES").font = Font(bold=True)
for c in COLS + [TOTC]:
    calc_cell(cs, f"{c}{r}", f"=SUM({c}{OPEX_START}:{c}{r-1})", '$#,##0', bold=True)
r += 2
section(cs, r, "CAPITAL & FINANCING", 14); r += 1
CAP_R = r
for name in ["Equipment purchases", "Loan drawdown (money in)", "Loan repayment (money out)",
             "Owner draw / dividends", "Owner injection (money in)"]:
    cs.cell(row=r, column=1, value=name).font = Font(bold=True)
    for c in COLS:
        input_cell(cs, f"{c}{r}", 0, '$#,##0')
    calc_cell(cs, f"{TOTC}{r}", f"=SUM(B{r}:M{r})", '$#,##0')
    r += 1
widths(cs, dict({"A": 34, TOTC: 14}, **{c: 12 for c in COLS}))
cs.sheet_view.showGridLines = False

# ---------- Cash Flow ----------
cf = wb.create_sheet("Cash Flow")
title_block(cf, "Cash Flow", "The tab that matters: when cash actually lands and leaves.", 14)
month_headers(cf, 4, "Cash movement")
r = 5
def line(label, builder, fmt='$#,##0', bold=False):
    global r
    cf.cell(row=r, column=1, value=label).font = Font(bold=bold)
    for i, c in enumerate(COLS):
        calc_cell(cf, f"{c}{r}", builder(i, c), fmt, bold=bold)
    calc_cell(cf, f"{TOTC}{r}", f"=SUM(B{r}:M{r})", fmt, bold=bold)
    out = r; r += 1
    return out

section(cf, 4, "", 14)
month_headers(cf, 4, "Cash movement")
OPEN_R = r
cf.cell(row=r, column=1, value="Opening cash").font = Font(bold=True)
for i, c in enumerate(COLS):
    calc_cell(cf, f"{c}{r}", f"={OPEN}", '$#,##0', bold=True)
r += 2
section(cf, r, "CASH IN", 14); r += 1
COLL_R = line("Collections from sales",
    lambda i, c: (f"=Revenue!{c}{REV_T}*{CIM}" if i == 0
                  else f"=Revenue!{c}{REV_T}*{CIM}+Revenue!{COLS[i-1]}{REV_T}*{C30}"))
HST_IN = line("Sales tax collected", lambda i, c: f"=Revenue!{c}{REV_T}*{HST}")
FIN_IN = line("Loan drawdown & owner injections", lambda i, c: f"=Costs!{c}{CAP_R+1}+Costs!{c}{CAP_R+4}")
IN_T = line("TOTAL CASH IN", lambda i, c: f"={c}{COLL_R}+{c}{HST_IN}+{c}{FIN_IN}", bold=True)
r += 1
section(cf, r, "CASH OUT", 14); r += 1
COGS_OUT = line("Payments to suppliers (COGS)",
    lambda i, c: (f"=Costs!{c}{COGS_R}*{SIM}" if i == 0
                  else f"=Costs!{c}{COGS_R}*{SIM}+Costs!{COLS[i-1]}{COGS_R}*{S30}"))
OPEX_OUT = line("Operating expenses", lambda i, c: f"=Costs!{c}{OPEX_T}")
HST_OUT = line("Sales tax remitted",
    lambda i, c: (f'=IF({HST}=0,0,IF(Assumptions!$B$15="Yes",IF(MOD({i+1},3)=0,SUM({get_column_letter(2+max(0,i-2))}{HST_IN}:{c}{HST_IN}),0),{c}{HST_IN}))'))
CAP_OUT = line("Equipment, loan repayments & draws",
    lambda i, c: f"=Costs!{c}{CAP_R}+Costs!{c}{CAP_R+2}+Costs!{c}{CAP_R+3}")
TAX_OUT = line("Income tax instalment",
    lambda i, c: f"=MAX(0,(Revenue!{c}{REV_T}-Costs!{c}{COGS_R}-Costs!{c}{OPEX_T}))*{TAXR}")
OUT_T = line("TOTAL CASH OUT", lambda i, c: f"={c}{COGS_OUT}+{c}{OPEX_OUT}+{c}{HST_OUT}+{c}{CAP_OUT}+{c}{TAX_OUT}", bold=True)
r += 1
NET_R = line("NET CASH MOVEMENT", lambda i, c: f"={c}{IN_T}-{c}{OUT_T}", bold=True)
CLOSE_R = line("CLOSING CASH", lambda i, c: f"={c}{OPEN_R}+{c}{NET_R}", bold=True)
# opening cash of each month after the first = previous month's closing cash
for i, c in enumerate(COLS):
    if i > 0:
        calc_cell(cf, f"{c}{OPEN_R}", f"={COLS[i-1]}{CLOSE_R}", '$#,##0', bold=True)
r += 1
FLAG_R = r
cf.cell(row=r, column=1, value="Cash warning").font = Font(bold=True, color=RED)
for c in COLS:
    calc_cell(cf, f"{c}{r}", f'=IF({c}{CLOSE_R}<0,"⚠ SHORT","OK")')
widths(cf, dict({"A": 36, TOTC: 14}, **{c: 12 for c in COLS}))
cf.sheet_view.showGridLines = False

# ---------- Dashboard ----------
d = wb.create_sheet("Dashboard")
title_block(d, "Dashboard", "Twelve months on one screen.", 6)
section(d, 4, "HEADLINES", 6)
rows = [("Total revenue", f"=Revenue!{TOTC}{REV_T}", '$#,##0'),
        ("Gross profit", f"=Revenue!{TOTC}{REV_T}-Costs!{TOTC}{COGS_R}", '$#,##0'),
        ("Gross margin", '=IF(B6=0,"—",B7/B6)', '0.0%'),
        ("Operating expenses", f"=Costs!{TOTC}{OPEX_T}", '$#,##0'),
        ("Net operating profit", "=B7-B9", '$#,##0'),
        ("Net margin", '=IF(B6=0,"—",B10/B6)', '0.0%'),
        ("Opening cash", f"={OPEN}", '$#,##0'),
        ("Closing cash after 12 months", f"=Cash Flow!{COLS[-1]}{CLOSE_R}", '$#,##0'),
        ("Lowest cash balance in the year", f"=MIN('Cash Flow'!B{CLOSE_R}:M{CLOSE_R})", '$#,##0'),
        ("Months with negative cash", f'=COUNTIF(\'Cash Flow\'!B{CLOSE_R}:M{CLOSE_R},"<0")', '0'),
        ("Monthly break-even revenue", f'=IF(1-{COGSPC}=0,"—",AVERAGE(Costs!B{OPEX_T}:M{OPEX_T})/(1-{COGSPC}))', '$#,##0')]
r = 6
for lab, f, fmt in rows:
    d.cell(row=r, column=1, value=lab).font = Font(bold=True)
    calc_cell(d, f"B{r}", f.replace("Cash Flow!", "'Cash Flow'!"), fmt, bold=True); r += 1
ch = LineChart(); ch.title = "Closing cash by month"; ch.height = 9; ch.width = 20
ch.add_data(Reference(cf, min_col=1, max_col=13, min_row=CLOSE_R, max_row=CLOSE_R), from_rows=True, titles_from_data=True)
d.add_chart(ch, "E6")
bar = BarChart(); bar.type = "col"; bar.title = "Cash in vs cash out"; bar.height = 9; bar.width = 20
bar.add_data(Reference(cf, min_col=1, max_col=13, min_row=IN_T, max_row=IN_T), from_rows=True, titles_from_data=True)
bar.add_data(Reference(cf, min_col=1, max_col=13, min_row=OUT_T, max_row=OUT_T), from_rows=True, titles_from_data=True)
d.add_chart(bar, "E26")
widths(d, {"A": 34, "B": 18})
d.sheet_view.showGridLines = False

start_here(wb, "Small Business 12-Month Cash Flow Forecast",
    "Profit is an opinion; cash is a fact. This shows you the fact, month by month.",
    ["Set your start month, opening cash and month-1 revenue on the Assumptions tab.",
     "Set your collection terms — how much you get paid in month, and how much 30 days later.",
     "Adjust the revenue streams and their growth on the Revenue tab.",
     "Type your real overheads on the Costs tab. Add capital spend and loan movements at the bottom.",
     "Read Cash Flow and the Dashboard: the lowest cash balance and any ⚠ SHORT months are what to act on."],
    ["Start Here — this page.",
     "Assumptions — one screen that drives the whole model, with built-in sanity checks.",
     "Revenue — four streams with growth and unit economics.",
     "Costs — COGS driven off revenue, 14 overhead lines, capital and financing.",
     "Cash Flow — timing-aware collections, supplier terms, HST remittance and tax instalments.",
     "Dashboard — margins, break-even revenue, lowest cash point and two charts."])
finish(wb, os.path.join(OUT, "05-Small-Business-12-Month-Cash-Flow-Forecast.xlsx"))
