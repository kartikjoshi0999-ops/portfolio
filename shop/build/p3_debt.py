import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference
from kjstyle import *

OUT = os.path.join(os.path.dirname(__file__), "..", "dist")
N = 8      # debt slots
M = 180    # months modelled
FIRST, LAST = 5, 12

wb = Workbook(); wb.remove(wb.active)

# ---------- Debts ----------
d = wb.create_sheet("Your Debts")
title_block(d, "Your Debts", "List up to 8 debts. Order does not matter — the plan sorts them for you.", 7)
header_row(d, 4, ["Debt name", "Balance owing", "Interest rate (APR)", "Minimum payment",
                  "Sort key", "Payoff order", "Position"])
sample = [("Visa credit card", 6400, 0.1999, 160), ("Store card", 1850, 0.2799, 60),
          ("Car loan", 14200, 0.0689, 385), ("Student loan", 9800, 0.0545, 210),
          ("Line of credit", 5200, 0.1145, 130)]
for k in range(N):
    r = FIRST + k
    nm, bal, apr, mn = sample[k] if k < len(sample) else ("", None, None, None)
    input_cell(d, f"A{r}", nm)
    input_cell(d, f"B{r}", bal, '$#,##0.00')
    input_cell(d, f"C{r}", apr, '0.00%')
    input_cell(d, f"D{r}", mn, '$#,##0.00')
    calc_cell(d, f"E{r}", f'=IF(OR($A{r}="",N($B{r})<=0),9999999,IF($B$16="Snowball",$B{r},1000000-N($C{r})*1000000))', '#,##0')
    calc_cell(d, f"F{r}", f'=IF(E{r}=9999999,"—",RANK.EQ(E{r},$E${FIRST}:$E${LAST},1)+COUNTIF($E${FIRST}:E{r},E{r})-1)', '0')
    calc_cell(d, f"G{r}", f'=IFERROR(MATCH({k+1},$F${FIRST}:$F${LAST},0),0)', '0')
r = LAST + 1
d.cell(row=r, column=1, value="Total owing").font = Font(bold=True)
calc_cell(d, f"B{r}", f"=SUM(B{FIRST}:B{LAST})", '$#,##0.00', bold=True)
d.cell(row=r, column=3, value="Total minimums").font = Font(bold=True)
calc_cell(d, f"D{r}", f"=SUM(D{FIRST}:D{LAST})", '$#,##0.00', bold=True)
r += 2  # row 15
section(d, r, "YOUR PLAN", 7); r += 1  # row 16
d.cell(row=r, column=1, value="Method").font = Font(bold=True)
input_cell(d, f"B{r}", "Snowball")
dv = DataValidation(type="list", formula1='"Snowball,Avalanche"'); d.add_data_validation(dv); dv.add(d[f"B{r}"])
d.cell(row=r, column=3, value="Snowball = smallest balance first (fastest wins). Avalanche = highest rate first (cheapest).").font = Font(size=9, italic=True, color=GREY)
r += 1  # 17
d.cell(row=r, column=1, value="Extra payment per month").font = Font(bold=True)
input_cell(d, f"B{r}", 250, '$#,##0.00')
r += 1  # 18
d.cell(row=r, column=1, value="Total monthly budget").font = Font(bold=True)
calc_cell(d, f"B{r}", f"=D{LAST+1}+B17", '$#,##0.00', bold=True)
r += 1  # 19
d.cell(row=r, column=1, value="First payment month").font = Font(bold=True)
input_cell(d, f"B{r}", "2026-01-01", None)
d[f"B{r}"].number_format = 'yyyy-mm'
r += 2
note(d, r, "Type the method exactly as Snowball or Avalanche — the drop-down keeps it clean.")
note(d, r + 1, "Extra payment is anything you can add on top of the minimums. Try changing it and watch the payoff date move.")
calc_cell(d, f"A{r+3}", f'=IF(B18>=D{LAST+1},"Budget covers every minimum payment.","\u26a0 Your budget is below the total minimums \u2014 raise it, or some accounts will go unpaid.")')
widths(d, {"A": 26, "B": 16, "C": 18, "D": 18, "E": 12, "F": 13, "G": 10})
d.sheet_view.showGridLines = False

BUDGET, START = "'Your Debts'!$B$18", "'Your Debts'!$B$19"

# ---------- Plan / schedule ----------
s = wb.create_sheet("Payoff Plan")
title_block(s, "Payoff Plan", "Every debt gets its minimum first. Everything spare goes to the target debt, and rolls on as each one clears.", 12)
s.cell(row=4, column=1, value="Payoff order \u2192").font = Font(bold=True)
for k in range(N):
    col = 6 + k * 4
    s.cell(row=4, column=col, value=f'=IF(\'Your Debts\'!$G${FIRST+k}=0,"\u2014",INDEX(\'Your Debts\'!$A${FIRST}:$A${LAST},\'Your Debts\'!$G${FIRST+k}))')
    s.cell(row=4, column=col).font = Font(bold=True, color=MID)
    s.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 3)
hdr = ["Month", "Date", "Budget", "Minimums due", "Spare for target"]
for k in range(N):
    hdr += ["Interest", "Minimum due", "Payment", "Balance"]
hdr += ["Total balance", "Interest paid to date"]
header_row(s, 5, hdr)
TOTCOL = 6 + N * 4
TL, IL = get_column_letter(TOTCOL), get_column_letter(TOTCOL + 1)
MIN_COLS = [get_column_letter(7 + k * 4) for k in range(N)]
PAY_COLS = [get_column_letter(8 + k * 4) for k in range(N)]
for m in range(M):
    r = 6 + m
    s.cell(row=r, column=1, value=m + 1).border = BOX
    calc_cell(s, f"B{r}", f"=EDATE({START},{m})", 'mmm yyyy')
    calc_cell(s, f"C{r}", f"={BUDGET}", '$#,##0.00')
    calc_cell(s, f"D{r}", "=" + "+".join(f"{c}{r}" for c in MIN_COLS), '$#,##0.00')
    calc_cell(s, f"E{r}", f"=MAX(0,C{r}-D{r})", '$#,##0.00')
    for k in range(N):
        col = 6 + k * 4
        ci, cm, cp, cb = (get_column_letter(col), get_column_letter(col + 1),
                          get_column_letter(col + 2), get_column_letter(col + 3))
        idx = f"'Your Debts'!$G${FIRST+k}"
        if m == 0:
            prev = f'IF({idx}=0,0,INDEX(\'Your Debts\'!$B${FIRST}:$B${LAST},{idx}))'
        else:
            prev = f"{cb}{r-1}"
        rate = f'IF({idx}=0,0,INDEX(\'Your Debts\'!$C${FIRST}:$C${LAST},{idx})/12)'
        minp = f'IF({idx}=0,0,INDEX(\'Your Debts\'!$D${FIRST}:$D${LAST},{idx}))'
        calc_cell(s, f"{ci}{r}", f"=ROUND(({prev})*({rate}),2)", '#,##0.00')
        calc_cell(s, f"{cm}{r}", f"=MIN(({prev})+{ci}{r},{minp})", '#,##0.00')
        spent = "0" if k == 0 else "+".join(f"({PAY_COLS[j]}{r}-{MIN_COLS[j]}{r})" for j in range(k))
        calc_cell(s, f"{cp}{r}",
                  f"={cm}{r}+MIN(({prev})+{ci}{r}-{cm}{r},MAX(0,$E{r}-({spent})))", '#,##0.00')
        calc_cell(s, f"{cb}{r}", f"=ROUND(MAX(0,({prev})+{ci}{r}-{cp}{r}),2)", '#,##0.00')
    calc_cell(s, f"{TL}{r}", "=" + "+".join(f"{get_column_letter(9 + k*4)}{r}" for k in range(N)),
              '$#,##0.00', bold=True)
    prev_i = "0" if m == 0 else f"{IL}{r-1}"
    calc_cell(s, f"{IL}{r}", f"={prev_i}+" + "+".join(f"{get_column_letter(6 + k*4)}{r}" for k in range(N)),
              '$#,##0.00')
widths(s, dict({"A": 8, "B": 12, "C": 12, "D": 14, "E": 15, TL: 16, IL: 20},
               **{get_column_letter(6 + i): 11 for i in range(N * 4)}))
s.freeze_panes = "F6"
s.sheet_view.showGridLines = False
note(s, 6 + M + 1, "Minimums due shrink as debts clear, so the spare column grows on its own \u2014 that is the snowball rolling forward.")

# ---------- Summary ----------
u = wb.create_sheet("Summary")
title_block(u, "Summary", "The two numbers that matter: when you are free, and what it cost.", 6)
section(u, 4, "YOUR RESULT", 6)
rows = [
    ("Method chosen", "='Your Debts'!B16", '@'),
    ("Total debt today", f"='Your Debts'!B{LAST+1}", '$#,##0.00'),
    ("Monthly budget", f"={BUDGET}", '$#,##0.00'),
    ("Months to debt-free", f'=IF(COUNTIF(\'Payoff Plan\'!${TL}$6:${TL}${5+M},">=0.01")={M},"Not paid off in 15 yrs",COUNTIF(\'Payoff Plan\'!${TL}$6:${TL}${5+M},">=0.01"))', '0'),
    ("Years to debt-free", '=IF(ISNUMBER(B9),B9/12,"—")', '0.0'),
    ("Debt-free date", f'=IF(ISNUMBER(B9),EDATE({START},B9-1),"—")', 'mmm yyyy'),
    ("Total interest you will pay", f"=INDEX('Payoff Plan'!${IL}$6:${IL}${5+M},IF(ISNUMBER(B9),B9,{M}))", '$#,##0.00'),
    ("Total you will repay", "=B7+B12", '$#,##0.00'),
]
r = 6
for lab, f, fmt in rows:
    u.cell(row=r, column=1, value=lab).font = Font(bold=True)
    calc_cell(u, f"B{r}", f, fmt, bold=True)
    r += 1
r += 1
section(u, r, "HOW TO USE THIS", 6); r += 1
for line in [
    "Change the extra payment on Your Debts and watch the debt-free date move. That is the whole exercise.",
    "Switch between Snowball and Avalanche to see the trade-off: Snowball clears accounts sooner and keeps you motivated; Avalanche costs less interest.",
    "Every time a debt hits zero, its payment rolls onto the next one automatically. Do not lower your budget when that happens — that is where the speed comes from.",
    "If the months-to-debt-free cell says not paid off, your budget is too close to the interest charge. Raise the extra payment or look at consolidating.",
]:
    c = u.cell(row=r, column=1, value="•  " + line)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    u.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    u.row_dimensions[r].height = 30
    r += 1
ch = LineChart(); ch.title = "Total balance over time"; ch.height = 9; ch.width = 20
ch.add_data(Reference(s, min_col=TOTCOL, min_row=5, max_row=5 + 120), titles_from_data=True)
u.add_chart(ch, "H6")
widths(u, {"A": 34, "B": 20})
u.sheet_view.showGridLines = False

start_here(wb, "Debt Payoff Planner — Snowball & Avalanche",
    "Put your debts in, pick a method, get a real payoff date.",
    ["List your debts on the Your Debts tab — name, balance, rate and minimum payment.",
     "Pick Snowball (smallest balance first) or Avalanche (highest rate first) from the drop-down.",
     "Enter any extra amount you can pay each month on top of the minimums.",
     "Read the Summary tab: your debt-free date and total interest.",
     "Open Payoff Plan for the month-by-month schedule you can actually follow."],
    ["Start Here — this page.",
     "Your Debts — up to 8 debts, with automatic payoff ordering.",
     "Payoff Plan — 180-month schedule with rollover payments built in.",
     "Summary — debt-free date, total interest, total repaid, and a balance chart."])
finish(wb, os.path.join(OUT, "03-Debt-Payoff-Planner-Snowball-and-Avalanche.xlsx"))
