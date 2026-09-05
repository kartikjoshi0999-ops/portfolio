"""Shared workbook styling for Joshi Finance Templates products."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = "0A1628"
MID = "0D1F3C"
GOLD = "C9A84C"
TEAL = "00B4A6"
LIGHT = "F4F6F9"
BAND = "EDF1F6"
GREY = "8A97AB"
RED = "C0392B"
GREEN = "1E8449"

BRAND = "Joshi Finance Templates"
AUTHOR = "Kartik Joshi, MBA — Financial Analyst & Banking Professional"

LICENSE = (
    "LICENSE: Single-user personal/business licence. You may use and edit this file for "
    "yourself or your own business. You may NOT resell, redistribute, share, or publish "
    "this file or any edited version of it. All rights reserved."
)
DISCLAIMER = (
    "DISCLAIMER: This template is an educational planning tool only. It is not financial, "
    "tax, investment, or legal advice. Figures such as contribution limits and tax rates "
    "change — always confirm current numbers with the CRA or a licensed advisor before "
    "acting on any output."
)

thin = Side(style="thin", color="C7D0DB")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


def title_block(ws, title, subtitle, width_cols=8):
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=10, color="C9A84C")
    for r in (1, 2):
        for c in range(1, width_cols + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18


def section(ws, row, text, width_cols=8):
    ws.cell(row=row, column=1, value=text).font = Font(size=11, bold=True, color="FFFFFF")
    for c in range(1, width_cols + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=MID)
    ws.row_dimensions[row].height = 20


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 28


def input_cell(ws, ref, value=None, numfmt=None):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.fill = PatternFill("solid", fgColor="FFF7DC")
    c.font = Font(color="1A1A1A", bold=True)
    c.border = BOX
    if numfmt:
        c.number_format = numfmt
    return c


def calc_cell(ws, ref, formula, numfmt=None, bold=False):
    c = ws[ref]
    c.value = formula
    c.fill = PatternFill("solid", fgColor=BAND)
    c.font = Font(bold=bold)
    c.border = BOX
    if numfmt:
        c.number_format = numfmt
    return c


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def note(ws, row, text, col=1, color=GREY):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(size=9, italic=True, color=color)
    return c


def money(ws, cell_range, fmt='#,##0.00'):
    for row in ws[cell_range]:
        for c in row:
            c.number_format = fmt


def start_here(wb, product, tagline, how_to, whats_inside):
    """Build the standard 'Start Here' tab and return it."""
    ws = wb.create_sheet("Start Here", 0)
    widths(ws, {"A": 4, "B": 100})
    title_block(ws, product, tagline, width_cols=3)
    r = 4
    ws.cell(row=r, column=2, value="WHAT'S INSIDE").font = Font(size=12, bold=True, color=MID)
    r += 1
    for line in whats_inside:
        ws.cell(row=r, column=2, value="•  " + line).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="HOW TO USE IT").font = Font(size=12, bold=True, color=MID)
    r += 1
    for i, line in enumerate(how_to, 1):
        ws.cell(row=r, column=2, value=f"{i}.  {line}").alignment = Alignment(wrap_text=True)
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="COLOUR KEY").font = Font(size=12, bold=True, color=MID)
    r += 1
    k = ws.cell(row=r, column=2, value="Yellow cells = you type here.")
    k.fill = PatternFill("solid", fgColor="FFF7DC")
    r += 1
    k = ws.cell(row=r, column=2, value="Grey cells = formulas. Do not overwrite them.")
    k.fill = PatternFill("solid", fgColor=BAND)
    r += 2
    for text in (LICENSE, DISCLAIMER):
        c = ws.cell(row=r, column=2, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=9, color=GREY)
        ws.row_dimensions[r].height = 46
        r += 1
    r += 1
    ws.cell(row=r, column=2, value=f"Created by {AUTHOR}").font = Font(size=10, bold=True, color=MID)
    r += 1
    ws.cell(row=r, column=2, value="Questions or a problem with the file? Message me through Etsy and I will help.").font = Font(size=9, color=GREY)
    ws.sheet_view.showGridLines = False
    return ws


def finish(wb, path):
    # force Excel / Sheets / LibreOffice to compute every formula the moment the file opens,
    # so a buyer sees a fully populated template on first open
    wb.calculation.fullCalcOnLoad = True
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = GOLD if ws.title == "Start Here" else TEAL
    wb.active = 0
    wb.save(path)
    print("wrote", path)
