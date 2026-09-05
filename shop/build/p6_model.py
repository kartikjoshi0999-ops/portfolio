import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from kjstyle import *

OUT = os.path.join(os.path.dirname(__file__), "..", "dist")
Y = ["C", "D", "E", "F", "G"]          # forecast years 1-5
B0 = "B"                                # historical / opening year
wb = Workbook(); wb.remove(wb.active)

def yr_header(ws, row, label="Line item"):
    header_row(ws, row, [label, "Year 0 (actual)", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])

def row_line(ws, r, label, b_formula, y_formula, fmt='$#,##0', bold=False, indent=0):
    c = ws.cell(row=r, column=1, value=("    " * indent) + label)
    c.font = Font(bold=bold)
    if b_formula is None:
        ws.cell(row=r, column=2, value="").border = BOX
    else:
        calc_cell(ws, f"B{r}", b_formula, fmt, bold=bold)
    for i, col in enumerate(Y):
        calc_cell(ws, f"{col}{r}", y_formula(i, col, Y[i-1] if i else "B"), fmt, bold=bold)
    return r

# ---------- Assumptions ----------
a = wb.create_sheet("Assumptions")
title_block(a, "Assumptions & Drivers", "Every yellow cell here drives all three statements.", 7)
section(a, 4, "OPERATING DRIVERS", 7)
yr_header(a, 5, "Driver")
drivers = [
    ("Revenue growth", None, [0.18, 0.15, 0.12, 0.10, 0.08], '0.0%'),
    ("Gross margin", 0.55, [0.56, 0.57, 0.57, 0.58, 0.58], '0.0%'),
    ("Operating expenses % of revenue", 0.34, [0.34, 0.33, 0.33, 0.32, 0.32], '0.0%'),
    ("Depreciation % of opening PP&E", 0.12, [0.12, 0.12, 0.12, 0.12, 0.12], '0.0%'),
    ("Capital expenditure % of revenue", 0.05, [0.06, 0.05, 0.05, 0.04, 0.04], '0.0%'),
    ("Interest rate on debt", 0.075, [0.075, 0.075, 0.075, 0.075, 0.075], '0.00%'),
    ("Effective tax rate", 0.26, [0.26, 0.26, 0.26, 0.26, 0.26], '0.0%'),
    ("Dividend payout % of net income", 0.0, [0.0, 0.10, 0.15, 0.20, 0.25], '0.0%'),
    ("Days sales outstanding (DSO)", 42, [40, 38, 36, 35, 35], '0'),
    ("Days inventory on hand (DIO)", 58, [55, 52, 50, 48, 46], '0'),
    ("Days payable outstanding (DPO)", 34, [36, 38, 40, 40, 40], '0'),
    ("Debt drawn / (repaid) in the year", 0, [0, 0, -250000, -250000, -250000], '$#,##0'),
]
r = 6
DRV = {}
for name, b, vals, fmt in drivers:
    a.cell(row=r, column=1, value=name).font = Font(bold=True)
    if b is None:
        a.cell(row=r, column=2, value="—").border = BOX
    else:
        input_cell(a, f"B{r}", b, fmt)
    for col, v in zip(Y, vals):
        input_cell(a, f"{col}{r}", v, fmt)
    DRV[name] = r
    r += 1
r += 1
section(a, r, "OPENING BALANCE SHEET (YEAR 0 ACTUALS)", 7); r += 1
OPEN = {}
openings = [("Revenue — Year 0", 4200000, '$#,##0'), ("Cash", 310000, '$#,##0'),
            ("Accounts receivable", 483000, '$#,##0'), ("Inventory", 300000, '$#,##0'),
            ("Property, plant & equipment (net)", 1450000, '$#,##0'),
            ("Accounts payable", 176000, '$#,##0'), ("Long-term debt", 900000, '$#,##0'),
            ("Share capital", 500000, '$#,##0'), ("Retained earnings", 967000, '$#,##0')]
for name, v, fmt in openings:
    a.cell(row=r, column=1, value=name).font = Font(bold=True)
    input_cell(a, f"B{r}", v, fmt)
    OPEN[name] = r
    r += 1
r += 1
calc_cell(a, f"B{r}", f'=IF(ROUND(B{OPEN["Cash"]}+B{OPEN["Accounts receivable"]}+B{OPEN["Inventory"]}+B{OPEN["Property, plant & equipment (net)"]}'
                      f'-B{OPEN["Accounts payable"]}-B{OPEN["Long-term debt"]}-B{OPEN["Share capital"]}-B{OPEN["Retained earnings"]},2)=0,'
                      f'"Opening balance sheet balances ✓","⚠ Opening balance sheet is out by "&TEXT(B{OPEN["Cash"]}+B{OPEN["Accounts receivable"]}+B{OPEN["Inventory"]}'
                      f'+B{OPEN["Property, plant & equipment (net)"]}-B{OPEN["Accounts payable"]}-B{OPEN["Long-term debt"]}-B{OPEN["Share capital"]}-B{OPEN["Retained earnings"]},"$#,##0"))')
a.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
widths(a, dict({"A": 40, "B": 18}, **{c: 14 for c in Y}))
a.sheet_view.showGridLines = False

A = lambda name: f"Assumptions!$A$1"  # placeholder
def dr(name, col):  return f"Assumptions!{col}${DRV[name]}"
def op(name):       return f"Assumptions!$B${OPEN[name]}"

# ---------- Income Statement ----------
i = wb.create_sheet("Income Statement")
title_block(i, "Income Statement", "Driven entirely from the Assumptions tab.", 7)
yr_header(i, 4)
REV = row_line(i, 5, "Revenue", f"={op('Revenue — Year 0')}",
               lambda k, c, p: f"={p}5*(1+{dr('Revenue growth', c)})", bold=True)
COGS = row_line(i, 6, "Cost of goods sold", f"=-B5*(1-{dr('Gross margin','$B')})",
                lambda k, c, p: f"=-{c}5*(1-{dr('Gross margin', c)})", indent=1)
GP = row_line(i, 7, "Gross profit", "=B5+B6", lambda k, c, p: f"={c}5+{c}6", bold=True)
OPEX = row_line(i, 8, "Operating expenses", f"=-B5*{dr('Operating expenses % of revenue','$B')}",
                lambda k, c, p: f"=-{c}5*{dr('Operating expenses % of revenue', c)}", indent=1)
EBITDA = row_line(i, 9, "EBITDA", "=B7+B8", lambda k, c, p: f"={c}7+{c}8", bold=True)
DA = row_line(i, 10, "Depreciation & amortisation", f"=-{op('Property, plant & equipment (net)')}*{dr('Depreciation % of opening PP&E','$B')}",
              lambda k, c, p: f"=-'Balance Sheet'!{p}10*{dr('Depreciation % of opening PP&E', c)}", indent=1)
EBIT = row_line(i, 11, "EBIT (operating profit)", "=B9+B10", lambda k, c, p: f"={c}9+{c}10", bold=True)
INT = row_line(i, 12, "Interest expense", f"=-{op('Long-term debt')}*{dr('Interest rate on debt','$B')}",
               lambda k, c, p: f"=-'Balance Sheet'!{p}16*{dr('Interest rate on debt', c)}", indent=1)
EBT = row_line(i, 13, "Profit before tax", "=B11+B12", lambda k, c, p: f"={c}11+{c}12", bold=True)
TAX = row_line(i, 14, "Income tax", f"=-MAX(0,B13)*{dr('Effective tax rate','$B')}",
               lambda k, c, p: f"=-MAX(0,{c}13)*{dr('Effective tax rate', c)}", indent=1)
NI = row_line(i, 15, "NET INCOME", "=B13+B14", lambda k, c, p: f"={c}13+{c}14", bold=True)
DIV = row_line(i, 17, "Dividends declared", "=0",
               lambda k, c, p: f"=-{c}15*{dr('Dividend payout % of net income', c)}")
row_line(i, 19, "Gross margin %", "=IFERROR(B7/B5,0)", lambda k, c, p: f"=IFERROR({c}7/{c}5,0)", '0.0%')
row_line(i, 20, "EBITDA margin %", "=IFERROR(B9/B5,0)", lambda k, c, p: f"=IFERROR({c}9/{c}5,0)", '0.0%')
row_line(i, 21, "Net margin %", "=IFERROR(B15/B5,0)", lambda k, c, p: f"=IFERROR({c}15/{c}5,0)", '0.0%')
widths(i, dict({"A": 34, "B": 16}, **{c: 15 for c in Y}))
i.sheet_view.showGridLines = False

# ---------- Balance Sheet ----------
b = wb.create_sheet("Balance Sheet")
title_block(b, "Balance Sheet", "Working capital is driven by DSO, DIO and DPO. Cash comes from the cash flow.", 7)
yr_header(b, 4)
section(b, 5, "ASSETS", 7)
CASH = row_line(b, 6, "Cash", f"={op('Cash')}", lambda k, c, p: f"={p}6+'Cash Flow'!{c}18")
AR = row_line(b, 7, "Accounts receivable", f"={op('Accounts receivable')}",
              lambda k, c, p: f"='Income Statement'!{c}5/365*{dr('Days sales outstanding (DSO)', c)}")
INV = row_line(b, 8, "Inventory", f"={op('Inventory')}",
               lambda k, c, p: f"=-'Income Statement'!{c}6/365*{dr('Days inventory on hand (DIO)', c)}")
CA = row_line(b, 9, "Total current assets", "=SUM(B6:B8)", lambda k, c, p: f"=SUM({c}6:{c}8)", bold=True)
PPE = row_line(b, 10, "Property, plant & equipment (net)", f"={op('Property, plant & equipment (net)')}",
               lambda k, c, p: f"={p}10+'Cash Flow'!{c}12*-1+'Income Statement'!{c}10")
TA = row_line(b, 11, "TOTAL ASSETS", "=B9+B10", lambda k, c, p: f"={c}9+{c}10", bold=True)
section(b, 13, "LIABILITIES & EQUITY", 7)
AP = row_line(b, 14, "Accounts payable", f"={op('Accounts payable')}",
              lambda k, c, p: f"=-'Income Statement'!{c}6/365*{dr('Days payable outstanding (DPO)', c)}")
CL = row_line(b, 15, "Total current liabilities", "=B14", lambda k, c, p: f"={c}14", bold=True)
DEBT = row_line(b, 16, "Long-term debt", f"={op('Long-term debt')}",
                lambda k, c, p: f"={p}16+{dr('Debt drawn / (repaid) in the year', c)}")
TL = row_line(b, 17, "Total liabilities", "=B15+B16", lambda k, c, p: f"={c}15+{c}16", bold=True)
SC = row_line(b, 18, "Share capital", f"={op('Share capital')}", lambda k, c, p: f"={p}18")
RE = row_line(b, 19, "Retained earnings", f"={op('Retained earnings')}",
              lambda k, c, p: f"={p}19+'Income Statement'!{c}15+'Income Statement'!{c}17")
TE = row_line(b, 20, "Total equity", "=B18+B19", lambda k, c, p: f"={c}18+{c}19", bold=True)
TLE = row_line(b, 21, "TOTAL LIABILITIES & EQUITY", "=B17+B20", lambda k, c, p: f"={c}17+{c}20", bold=True)
CHK = row_line(b, 23, "CHECK — assets less liabilities & equity", "=ROUND(B11-B21,2)",
               lambda k, c, p: f"=ROUND({c}11-{c}21,2)", bold=True)
b.cell(row=24, column=1, value="Status").font = Font(bold=True)
for col in ["B"] + Y:
    calc_cell(b, f"{col}24", f'=IF(ABS({col}23)<0.5,"BALANCED ✓","⚠ OUT")')
widths(b, dict({"A": 38, "B": 16}, **{c: 15 for c in Y}))
b.sheet_view.showGridLines = False

# ---------- Cash Flow ----------
c_ = wb.create_sheet("Cash Flow")
title_block(c_, "Cash Flow Statement", "Indirect method. Ties back to the cash line on the balance sheet.", 7)
yr_header(c_, 4)
section(c_, 5, "OPERATING", 7)
row_line(c_, 6, "Net income", None, lambda k, c, p: f"='Income Statement'!{c}15")
row_line(c_, 7, "Add back depreciation & amortisation", None, lambda k, c, p: f"=-'Income Statement'!{c}10")
row_line(c_, 8, "(Increase) / decrease in receivables", None, lambda k, c, p: f"='Balance Sheet'!{p}7-'Balance Sheet'!{c}7")
row_line(c_, 9, "(Increase) / decrease in inventory", None, lambda k, c, p: f"='Balance Sheet'!{p}8-'Balance Sheet'!{c}8")
row_line(c_, 10, "Increase / (decrease) in payables", None, lambda k, c, p: f"='Balance Sheet'!{c}14-'Balance Sheet'!{p}14")
row_line(c_, 11, "Cash from operations", None, lambda k, c, p: f"=SUM({c}6:{c}10)", bold=True)
section(c_, 12 - 12 + 12, "", 7)
c_.cell(row=12, column=1, value="Capital expenditure").font = Font(bold=True)
for k, col in enumerate(Y):
    calc_cell(c_, f"{col}12", f"=-'Income Statement'!{col}5*{dr('Capital expenditure % of revenue', col)}", '$#,##0')
row_line(c_, 13, "Cash used in investing", None, lambda k, c, p: f"={c}12", bold=True)
row_line(c_, 15, "Debt drawn / (repaid)", None, lambda k, c, p: f"={dr('Debt drawn / (repaid) in the year', c)}")
row_line(c_, 16, "Dividends paid", None, lambda k, c, p: f"='Income Statement'!{c}17")
row_line(c_, 17, "Cash from financing", None, lambda k, c, p: f"={c}15+{c}16", bold=True)
row_line(c_, 18, "NET CHANGE IN CASH", None, lambda k, c, p: f"={c}11+{c}13+{c}17", bold=True)
row_line(c_, 19, "Opening cash", None, lambda k, c, p: f"='Balance Sheet'!{p}6")
row_line(c_, 20, "CLOSING CASH", None, lambda k, c, p: f"={c}19+{c}18", bold=True)
row_line(c_, 22, "Free cash flow (operations less capex)", None, lambda k, c, p: f"={c}11+{c}12", bold=True)
widths(c_, dict({"A": 40, "B": 16}, **{c: 15 for c in Y}))
c_.sheet_view.showGridLines = False

# ---------- Ratios ----------
q = wb.create_sheet("Ratios & Dashboard")
title_block(q, "Ratios & Dashboard", "The numbers a credit officer or investor asks for first.", 7)
yr_header(q, 4)
ratios = [
    ("Revenue growth", lambda k, c, p: f"=IFERROR('Income Statement'!{c}5/'Income Statement'!{p}5-1,0)", '0.0%'),
    ("Gross margin", lambda k, c, p: f"=IFERROR('Income Statement'!{c}7/'Income Statement'!{c}5,0)", '0.0%'),
    ("EBITDA margin", lambda k, c, p: f"=IFERROR('Income Statement'!{c}9/'Income Statement'!{c}5,0)", '0.0%'),
    ("Net margin", lambda k, c, p: f"=IFERROR('Income Statement'!{c}15/'Income Statement'!{c}5,0)", '0.0%'),
    ("Current ratio", lambda k, c, p: f"=IFERROR('Balance Sheet'!{c}9/'Balance Sheet'!{c}15,0)", '0.00'),
    ("Quick ratio", lambda k, c, p: f"=IFERROR(('Balance Sheet'!{c}6+'Balance Sheet'!{c}7)/'Balance Sheet'!{c}15,0)", '0.00'),
    ("Debt / equity", lambda k, c, p: f"=IFERROR('Balance Sheet'!{c}16/'Balance Sheet'!{c}20,0)", '0.00'),
    ("Debt / EBITDA", lambda k, c, p: f"=IFERROR('Balance Sheet'!{c}16/'Income Statement'!{c}9,0)", '0.00'),
    ("Interest coverage (EBIT / interest)", lambda k, c, p: f"=IFERROR('Income Statement'!{c}11/-'Income Statement'!{c}12,0)", '0.00'),
    ("Return on equity", lambda k, c, p: f"=IFERROR('Income Statement'!{c}15/'Balance Sheet'!{c}20,0)", '0.0%'),
    ("Return on assets", lambda k, c, p: f"=IFERROR('Income Statement'!{c}15/'Balance Sheet'!{c}11,0)", '0.0%'),
    ("Cash conversion cycle (days)", lambda k, c, p: f"={dr('Days sales outstanding (DSO)', c)}+{dr('Days inventory on hand (DIO)', c)}-{dr('Days payable outstanding (DPO)', c)}", '0'),
    ("Free cash flow", lambda k, c, p: f"='Cash Flow'!{c}22", '$#,##0'),
]
r = 5
for name, fn, fmt in ratios:
    q.cell(row=r, column=1, value=name).font = Font(bold=True)
    q.cell(row=r, column=2, value="—").border = BOX
    for k, col in enumerate(Y):
        calc_cell(q, f"{col}{r}", fn(k, col, Y[k-1] if k else "B"), fmt)
    r += 1
r += 1
section(q, r, "MODEL INTEGRITY", 7); r += 1
q.cell(row=r, column=1, value="Balance sheet check").font = Font(bold=True)
for col in Y:
    calc_cell(q, f"{col}{r}", f"='Balance Sheet'!{col}24")
r += 1
q.cell(row=r, column=1, value="Cash flow ties to balance sheet").font = Font(bold=True)
for col in Y:
    calc_cell(q, f"{col}{r}", f'=IF(ABS(\'Cash Flow\'!{col}20-\'Balance Sheet\'!{col}6)<0.5,"TIES ✓","⚠ OUT")')
ch = BarChart(); ch.type = "col"; ch.title = "Revenue and EBITDA"; ch.height = 9; ch.width = 18
ch.add_data(Reference(i, min_col=1, max_col=7, min_row=5, max_row=5), from_rows=True, titles_from_data=True)
ch.add_data(Reference(i, min_col=1, max_col=7, min_row=9, max_row=9), from_rows=True, titles_from_data=True)
q.add_chart(ch, "I5")
widths(q, dict({"A": 36, "B": 14}, **{c: 14 for c in Y}))
q.sheet_view.showGridLines = False

start_here(wb, "3-Statement Financial Model Template",
    "Income statement, balance sheet and cash flow — linked, checked, and built the way analysts build them.",
    ["Put your Year 0 actuals into the opening balance sheet block on the Assumptions tab.",
     "Set the operating drivers: growth, margins, working-capital days, capex, tax and debt.",
     "Read the three statements. You never type into them — they are entirely formula-driven.",
     "Check the two integrity flags on Ratios & Dashboard: the balance sheet must say BALANCED, cash must say TIES.",
     "Flex the drivers to build your base, upside and downside cases."],
    ["Start Here — this page.",
     "Assumptions & Drivers — 12 drivers across 5 years plus an opening balance sheet with a balance check.",
     "Income Statement — revenue to net income with margin lines.",
     "Balance Sheet — working capital from DSO/DIO/DPO, PP&E roll-forward, retained earnings roll.",
     "Cash Flow — indirect method, tying back to the balance sheet cash line.",
     "Ratios & Dashboard — 13 analyst ratios, two integrity checks and a chart."])
finish(wb, os.path.join(OUT, "06-3-Statement-Financial-Model-Template.xlsx"))
